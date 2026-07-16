#!/usr/bin/env python3
"""Daily KUB holder summary — appends ONE row/day to the central sheet.

Fetches all EOA holders (native KUB >= 1) + KKUB holders from kubscan,
merges them, EXCLUDES CEX (Bitkub Exchange hot+cold) + Ecosystem Fund
(SBDF + Ecosystem Fund) + KUB Company Reserve, then stores only the
summary metrics (NOT per-wallet data):

  Date · EOA holders >=1 · Total held · native KUB · KKUB · Median ·
  Average · Top 10 share % · Top 100 share %

Idempotent per day: re-running the same date overwrites that row.
Outputs daily_summary.csv (canonical log) + KUB_Holders_Daily.xlsx (viewer).
"""
import csv
import json
import os
import re
import time
import urllib.parse
import urllib.request
from datetime import datetime, timezone, timedelta

DIR = os.path.dirname(os.path.abspath(__file__))
CSV_F = os.path.join(DIR, "daily_summary.csv")
XLSX_F = os.path.join(DIR, "KUB_Holders_Daily.xlsx")
JS_F = os.path.join(DIR, "daily_data.js")
LOG_F = os.path.join(DIR, "daily_summary.log")
KKUB = "0x67eBD850304c70d983B2d1b93ea79c7CD6c3F6b5"
HEADERS = {"User-Agent": "Mozilla/5.0 (bbt-bd-sosmart daily-summary)"}
MIN_WEI = 10**18
BURN = "0x0000000000000000000000000000000000000000"

CEX_RE = re.compile(r"Bitkub Exchange", re.I)
FUND_RE = re.compile(r"Sustainable Blockchain Development Fund|Ecosystem Fund", re.I)
RESERVE_RE = re.compile(r"KUB Company Reserve", re.I)

# Hardcoded exclusion set (CEX hot+cold / SBDF / Ecosystem Fund / KUB Company
# Reserve). Snapshot taken 2026-07-14 from kubscan public_tags. This is the
# authoritative fallback: kubscan's public_tags service is flaky and has
# intermittently returned EMPTY tags for these institutional wallets, which
# caused ~14M reserve KUB to leak into the "public holder" totals on 2026-07-14.
# We exclude by address (tag-independent) AND still honor live tags, so any
# newly-tagged wallet is also caught. Refresh this list if new fund/CEX/reserve
# wallets appear.
EXCLUDE_ADDRS = frozenset({
    # KUB: Bitkub Exchange (hot + cold wallets)
    "0xd4d189ae4c76dae3da202d285f86dced8dcf7f4a",
    "0x275f6bf5e301b56c5bdb756b0caa076fdcce3294",
    "0x593f38e99fe02600832f4905234dfeef76aa169c",
    "0x6ce8e06bf6c91417fd3d73fe5d835b343063b3e2",
    "0x71cacef3a37379abaee32a9cc2474176b7edef98",
    "0x7368735e1500ff6b63f0ba34cb436b8a0d9a38d6",
    "0x941a98bb32dce53bc8da4f453e662ffba3f8c91d",
    "0xa0cc55b76d53b9dc324006f8d57c4014e4fafcde",
    "0xae3217ec059bc2252f06fd757f4e593f25aeb5d8",
    "0xb9c3d85a1f9362c799216e608b2913510d5d184e",
    "0xbd8b12af54d1ffc09452a0b2a8ade0cc1d15fa66",
    "0xd8008a8c32950046a6a0e2e820adcd536dde695f",
    "0x96f00c26d3eb5865bd1a32eb1a67f65c3e1e2797",
    "0x08372d5b86aabc40ec86981e43e4786cc906114e",
    "0x545dec78b30f1ebfb62bdd38b304b87a82a87519",
    "0x5f001f54ec8fbd3c9f7d21c6a5a3302001144cc1",
    "0x7081aa230336d50b94909b617a236308c02d2d05",
    "0xfc78c1d8755aa70bc647700751938139f8708979",
    "0x08134b2966d7e5618d9b5a1f93ad8870bb087e0a",
    "0x2a9dd53872184b0b388871781e1c534b3f8a6f5c",
    "0x35fdd152014c308cc366e45196d5ca55eb9bf5aa",
    "0x5f6174de4883507914ecd147c00e35731963a287",
    "0x6066ea738a706304af4d190fc8148dfe2adc0451",
    "0xa8e15e6932ad47970d0c6cdd64b38409d15ab80a",
    "0xcae02305a617f02cfed1eee510e0aacba64dfd85",
    "0xd51977de87d61935f0a0ae0df21f0038ab5ba66b",
    "0xf772f8c33f4f771af1557989f4d5f14a066c6b37",
    "0xae2c036d891156c9423ebf40f696ca5f11f4a7cb",
    # Sustainable Blockchain Development Fund (hot + cold wallets)
    "0xf9efec93303ec5ee1d7c77497dd71f642eebebe3",
    "0x0aeb9606595abb5487d8e0eee53d458606fdbef5",
    "0x18e6a937ac9ca9fae9c6f66a0548fd500f6f8a4c",
    "0x4851cc948a07b4e19cb650ec81d34352f9514572",
    "0x731c6b625c13f913812c249a5ee71f0ccced687c",
    "0x88c371995c92109e34f3707e172da60aa62318c6",
    "0xc8b8011b545f750252184a3c03fe29f2094faf53",
    "0xd80f34e48ca2da7cf182037231d56d78e13a889a",
    "0x99221d359e17af05dda830e3770e06bc8215c203",
    "0xba0f259e78419ab996c537e5071dceec30ac243b",
    "0xc4772ee9146384efc05a3db91553c05ff4ac1c34",
    "0x8a2bf2c3af1fa2024d99080b3402f5d56ba6ab4c",
    # Ecosystem Fund (hot + cold wallets)
    "0xde1c5de071836fbfe75e0c981277f0b0bc0018ac",
    "0xd33883bd1c1ccfd1c1201934322a05f16a276bd0",
    "0x9d7faa9ff6e318b0702f53d6662860b6f0fc9bbb",
    "0x586cb64aa72cfa3fdf6b87116ae00f1c50ea02a3",
    # KUB Company Reserve (cold wallets)
    "0xacab0935deb68a23c0bafefbed19df1ab24a14c4",
    "0xfeab40137e5d4db7a7cc62d71ea87ab0f9e66329",
})

# Alert if native day-over-day swing exceeds this fraction (leak/tag-dropout guard)
NATIVE_JUMP_GUARD = 0.20

COLS = ["date", "eoa_holders", "total_kub", "native_kub", "kkub",
        "median_kub", "average_kub", "top10_share_pct", "top100_share_pct"]


def log(msg):
    line = f"{datetime.now(timezone.utc).isoformat(timespec='seconds')}  {msg}"
    print(line, flush=True)
    with open(LOG_F, "a") as f:
        f.write(line + "\n")


def get(url, retries=6):
    for i in range(retries):
        try:
            req = urllib.request.Request(url, headers=HEADERS)
            with urllib.request.urlopen(req, timeout=30) as r:
                return json.load(r)
        except Exception as e:
            wait = min(2 ** i, 30)
            log(f"  retry {i+1}/{retries}: {e} (sleep {wait}s)")
            time.sleep(wait)
    raise RuntimeError(f"gave up on {url}")


def fetch_native():
    """EOA native holders >= 1 KUB -> {addr_lower: (balance, tag)}."""
    out, nxt, pages = {}, None, 0
    base = "https://www.kubscan.com/api/v2/addresses"
    while True:
        url = base + ("?" + urllib.parse.urlencode(nxt) if nxt else "")
        data = get(url)
        stop = False
        for it in data.get("items", []):
            bal = int(it["coin_balance"] or 0)
            if bal < MIN_WEI:
                stop = True
                break
            if it["is_contract"] or it["hash"].lower() == BURN:
                continue
            tags = [t["display_name"] for t in (it.get("public_tags") or [])]
            out[it["hash"].lower()] = (bal / 1e18, tags[0] if tags else None)
        pages += 1
        nxt = data.get("next_page_params")
        if stop or not nxt:
            break
        time.sleep(0.05)
    log(f"native: {len(out)} EOA holders (>=1 KUB) in {pages} pages")
    return out


def fetch_kkub():
    """EOA KKUB holders -> {addr_lower: balance}."""
    out, nxt, pages = {}, None, 0
    base = f"https://www.kubscan.com/api/v2/tokens/{KKUB}/holders"
    while True:
        url = base + ("?" + urllib.parse.urlencode(nxt) if nxt else "")
        data = get(url)
        for it in data.get("items", []):
            if it["address"]["is_contract"]:
                continue
            out[it["address"]["hash"].lower()] = out.get(
                it["address"]["hash"].lower(), 0) + int(it["value"]) / 1e18
        pages += 1
        nxt = data.get("next_page_params")
        if not nxt:
            break
        time.sleep(0.05)
    log(f"kkub: {len(out)} EOA holders in {pages} pages")
    return out


def excluded(addr, tag):
    # Address-based exclusion is tag-independent, so a transient empty-tag
    # response from kubscan can no longer leak reserve/CEX/fund balances.
    if addr in EXCLUDE_ADDRS:
        return True
    return bool(tag) and (CEX_RE.search(tag) or FUND_RE.search(tag) or RESERVE_RE.search(tag))


def compute():
    native = fetch_native()
    kkub = fetch_kkub()
    # merge -> list of (total, native, kkub) after exclusions
    rows = []
    for a, (n, tag) in native.items():
        k = kkub.pop(a, 0)  # pop first so excluded wallets don't leak KKUB below
        if excluded(a, tag):
            continue
        rows.append((n + k, n, k))
    for a, k in kkub.items():  # KKUB-only EOAs (little/no native)
        if a in EXCLUDE_ADDRS:
            continue
        if k >= 1:
            rows.append((k, 0.0, k))
    rows = [r for r in rows if r[0] >= 1]
    rows.sort(key=lambda r: -r[0])

    n = len(rows)
    total = sum(r[0] for r in rows)
    native_tot = sum(r[1] for r in rows)
    kkub_tot = sum(r[2] for r in rows)
    median = rows[n // 2][0] if n else 0
    avg = total / n if n else 0
    top10 = sum(r[0] for r in rows[:10]) / total * 100 if total else 0
    top100 = sum(r[0] for r in rows[:100]) / total * 100 if total else 0
    return {
        "eoa_holders": n,
        "total_kub": round(total, 2),
        "native_kub": round(native_tot, 2),
        "kkub": round(kkub_tot, 2),
        "median_kub": round(median, 4),
        "average_kub": round(avg, 2),
        "top10_share_pct": round(top10, 2),
        "top100_share_pct": round(top100, 2),
    }


def upsert_csv(date_str, stats):
    rows = {}
    if os.path.exists(CSV_F):
        with open(CSV_F) as f:
            for r in csv.DictReader(f):
                rows[r["date"]] = r
    rows[date_str] = {"date": date_str, **{k: stats[k] for k in COLS[1:]}}
    ordered = [rows[d] for d in sorted(rows)]
    with open(CSV_F, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=COLS)
        w.writeheader()
        w.writerows(ordered)
    return ordered


def build_xlsx(rows):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Daily"
    titles = ["Date", "EOA holders ≥1", "Total held (KUB)", "Native KUB", "KKUB",
              "Median (KUB)", "Average (KUB)", "Top 10 share", "Top 100 share"]
    ws.append(["KUB Chain — EOA Holders Daily Tracker (filtered: excl. CEX / Ecosystem Fund / KUB Company Reserve)"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(titles))
    ws["A1"].font = Font(bold=True, size=13, color="FF2FD27D")
    ws.append(titles)
    hdr_fill = PatternFill("solid", fgColor="FF14532D")
    thin = Side(style="thin", color="FF2A3A2E")
    for c in ws[2]:
        c.font = Font(bold=True, color="FFFFFFFF")
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = Border(bottom=thin)
    fmt = {"total_kub": "#,##0", "native_kub": "#,##0", "kkub": "#,##0",
           "eoa_holders": "#,##0", "median_kub": "#,##0.00", "average_kub": "#,##0.00",
           "top10_share_pct": '0.00"%"', "top100_share_pct": '0.00"%"'}
    for r in rows:
        ws.append([
            r["date"], int(r["eoa_holders"]), float(r["total_kub"]), float(r["native_kub"]),
            float(r["kkub"]), float(r["median_kub"]), float(r["average_kub"]),
            float(r["top10_share_pct"]), float(r["top100_share_pct"]),
        ])
        row = ws[ws.max_row]
        for i, key in enumerate(COLS):
            if key in fmt:
                row[i].number_format = fmt[key]
    widths = [12, 14, 16, 14, 14, 13, 14, 13, 13]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A3"
    wb.save(XLSX_F)


def build_js(rows):
    """Emit daily_data.js so the HTML dashboard loads via <script> (works over file://)."""
    typed = []
    for r in rows:
        typed.append({
            "date": r["date"],
            "eoa_holders": int(r["eoa_holders"]),
            "total_kub": float(r["total_kub"]),
            "native_kub": float(r["native_kub"]),
            "kkub": float(r["kkub"]),
            "median_kub": float(r["median_kub"]),
            "average_kub": float(r["average_kub"]),
            "top10_share_pct": float(r["top10_share_pct"]),
            "top100_share_pct": float(r["top100_share_pct"]),
        })
    with open(JS_F, "w") as f:
        f.write("window.DAILY_DATA = " + json.dumps(typed, separators=(",", ":")) + ";\n")


def prev_native(date_str):
    """Most recent native_kub on a date strictly before date_str, or None."""
    if not os.path.exists(CSV_F):
        return None
    prev = None
    with open(CSV_F) as f:
        for r in csv.DictReader(f):
            if r["date"] < date_str and r.get("native_kub"):
                prev = float(r["native_kub"])  # rows are date-sorted -> last wins
    return prev


def main():
    # Bangkok date (UTC+7) so "today" matches the user's calendar day
    date_str = (datetime.now(timezone.utc) + timedelta(hours=7)).strftime("%Y-%m-%d")
    log(f"=== daily summary for {date_str} (Asia/Bangkok) ===")
    stats = compute()

    # Sanity guard: a huge native swing usually means kubscan dropped tags for
    # institutional wallets. Re-fetch once; if it persists, abort rather than
    # write a corrupted row (a scheduled run will report the failure).
    pv = prev_native(date_str)
    if pv and pv > 0:
        dev = abs(stats["native_kub"] - pv) / pv
        if dev > NATIVE_JUMP_GUARD:
            log(f"  WARNING native swing {dev*100:.1f}% vs prev {pv:,.0f} "
                f"(new {stats['native_kub']:,.0f}) -> re-fetching to confirm")
            stats = compute()
            dev = abs(stats["native_kub"] - pv) / pv
            if dev > NATIVE_JUMP_GUARD:
                raise RuntimeError(
                    f"native_kub swing {dev*100:.1f}% vs previous day persists "
                    f"across two fetches (prev {pv:,.0f}, new {stats['native_kub']:,.0f}). "
                    f"Likely a kubscan tag dropout or a new untagged reserve/CEX wallet. "
                    f"Row NOT written; check EXCLUDE_ADDRS / top holders before retrying.")
            log("  re-fetch within tolerance; proceeding")

    log("stats: " + json.dumps(stats))
    rows = upsert_csv(date_str, stats)
    build_xlsx(rows)
    build_js(rows)
    log(f"wrote {len(rows)} day(s) -> {os.path.basename(CSV_F)} + {os.path.basename(XLSX_F)} + {os.path.basename(JS_F)}")


if __name__ == "__main__":
    main()
